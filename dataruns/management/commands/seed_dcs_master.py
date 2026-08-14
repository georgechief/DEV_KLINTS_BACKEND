from __future__ import annotations

import re
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from dataruns.models import CheckMaster, DimensionMaster, RootCauseMaster

# Keep in sync with FE-03 / CheckMaster.is_optional seed policy.
OPTIONAL_CHECK_IDS = frozenset({"FD-03"})

try:
    from openpyxl import load_workbook
except ImportError as exc:
    load_workbook = None
    _OPENPYXL_IMPORT_ERROR = exc
else:
    _OPENPYXL_IMPORT_ERROR = None

WORKBOOK_RELATIVE_PATH = Path(
    "docs/dcs_scoring/Klints_Spec_InitialDataConsistencyCheck_v1.4.1_20260718.xlsx"
)
SHEET_DIMENSIONS = "07 Scoring Model"
SHEET_ROOT_CAUSES = "03 Root Cause Taxonomy"
SHEET_CHECKS = "09 MVP1 Check Scope"
SHEET_CATALOGUE = "02 Check Catalogue"

DIMENSION_DATA_START_ROW = 6
ROOT_CAUSE_DATA_START_ROW = 6
CHECK_DATA_START_ROW = 6
CATALOGUE_DATA_START_ROW = 6

DIMENSION_LABEL_RE = re.compile(r"^(\d{2})\s+(.+)$")
ROOT_CAUSE_CODE_RE = re.compile(r"RC-\d{2}")
PERCENT_IN_RULE_RE = re.compile(r"(\d+)\s*%")
RESULT_STATUS_HEADER = "Result status"
CONFIDENCE_HEADER = "Confidence"
FINAL_STATE_HEADER = "Final state"


class Command(BaseCommand):
    help = "Seed DCS master tables from the specification workbook."

    def handle(self, *args, **options):
        if load_workbook is None:
            raise CommandError(
                "openpyxl is required to read the DCS workbook. "
                "Install it with: pip install openpyxl"
            ) from _OPENPYXL_IMPORT_ERROR

        workbook_path = Path(settings.BASE_DIR) / WORKBOOK_RELATIVE_PATH
        if not workbook_path.exists():
            raise CommandError(f"Workbook not found: {workbook_path}")

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            with transaction.atomic():
                scoring_json = self._load_scoring_model_json(workbook)
                dimension_counts = self._seed_dimensions(workbook, scoring_json)
                root_cause_counts = self._seed_root_causes(workbook)
                catalogue_by_check_id = self._load_catalogue_rows(workbook)
                check_counts = self._seed_checks(workbook, catalogue_by_check_id)
        finally:
            workbook.close()

        self.stdout.write(
            self.style.SUCCESS(
                "DCS master seed complete:\n"
                f"  Dimensions inserted: {dimension_counts['created']}\n"
                f"  Root causes inserted: {root_cause_counts['created']}\n"
                f"  Checks inserted: {check_counts['created']}"
            )
        )
        try:
            from dataruns.dcs.gates import clear_optional_check_ids_cache
            from dataruns.dcs.master import clear_check_master_cache

            clear_check_master_cache()
            clear_optional_check_ids_cache()
        except Exception:  # noqa: BLE001
            pass

    def _load_scoring_model_json(self, workbook) -> dict:
        worksheet = workbook[SHEET_DIMENSIONS]
        result_status_json: dict[str, dict] = {}
        confidence_json: dict[str, dict] = {}
        final_state_json: list[dict] = []
        section: str | None = None

        for row in worksheet.iter_rows(values_only=True):
            header = self._cell_text(row, 0)
            if header == RESULT_STATUS_HEADER:
                section = "result_status"
                continue
            if header == CONFIDENCE_HEADER:
                section = "confidence"
                continue
            if header == FINAL_STATE_HEADER:
                section = "final_state"
                continue

            if section == "result_status":
                status = header
                if status in ("Result status", ""):
                    continue
                score_factor = row[1] if len(row) > 1 else None
                meaning = self._cell_text(row, 2)
                if not status:
                    section = None
                    continue
                result_status_json[status] = {
                    "score_factor": self._to_optional_number(score_factor),
                    "meaning": meaning,
                }
                continue

            if section == "confidence":
                level = header
                if level in ("Confidence", ""):
                    continue
                numeric_factor = row[1] if len(row) > 1 else None
                description = self._cell_text(row, 2)
                if not level:
                    section = None
                    continue
                confidence_json[level] = {
                    "numeric_factor": self._to_optional_number(numeric_factor),
                    "description": description,
                }
                continue

            if section == "final_state":
                final_state = header
                if final_state in ("Final state", ""):
                    continue
                condition = self._cell_text(row, 1)
                effect = self._cell_text(row, 2)
                if not final_state:
                    section = None
                    continue
                final_state_json.append(
                    {
                        "final_state": final_state,
                        "condition": condition,
                        "effect": effect,
                    }
                )

        if not result_status_json:
            raise CommandError(
                f"Could not load result status values from {SHEET_DIMENSIONS}"
            )
        if not confidence_json:
            raise CommandError(
                f"Could not load confidence values from {SHEET_DIMENSIONS}"
            )
        if not final_state_json:
            raise CommandError(
                f"Could not load final state values from {SHEET_DIMENSIONS}"
            )

        return {
            "result_status_json": result_status_json,
            "confidence_json": confidence_json,
            "final_state_json": final_state_json,
        }

    def _seed_dimensions(self, workbook, scoring_json: dict) -> dict[str, int]:
        worksheet = workbook[SHEET_DIMENSIONS]
        dimension_rows_by_id: dict[str, dict] = {}

        for row in worksheet.iter_rows(
            min_row=DIMENSION_DATA_START_ROW,
            values_only=True,
        ):
            label = self._cell_text(row, 0)
            if not label:
                continue

            parsed = self._parse_dimension_label(label)
            if parsed is None:
                continue

            dimension_id, name = parsed
            dimension_rows_by_id[dimension_id] = {
                "key": label,
                "name": name,
                "purpose": self._cell_text(row, 3),
                "percent_needed": self._parse_percent_from_rule(self._cell_text(row, 4)),
                "weight_percent": self._to_int(row[1], default=0),
            }

        dimension_labels = self._collect_check_dimension_labels(workbook)
        created = 0

        for label in sorted(dimension_labels):
            parsed = self._parse_dimension_label(label)
            if parsed is None:
                raise CommandError(f"Unrecognized dimension label: {label!r}")

            dimension_id, name = parsed
            workbook_row = dimension_rows_by_id.get(dimension_id)
            if workbook_row is None:
                self.stdout.write(
                    self.style.WARNING(
                        f"Dimension {label!r} is not present in {SHEET_DIMENSIONS}; "
                        "purpose and percent_needed are not available in the workbook."
                    )
                )
                defaults = {
                    "key": label,
                    "name": name,
                    "purpose": "",
                    "percent_needed": None,
                    "weight_percent": 0,
                }
            else:
                defaults = {
                    "key": workbook_row["key"],
                    "name": workbook_row["name"],
                    "purpose": workbook_row["purpose"],
                    "percent_needed": workbook_row["percent_needed"],
                    "weight_percent": workbook_row["weight_percent"],
                }

            defaults.update(
                {
                    "result_status_json": scoring_json["result_status_json"],
                    "confidence_json": scoring_json["confidence_json"],
                    "final_state_json": scoring_json["final_state_json"],
                    "is_active": True,
                }
            )

            _, was_created = DimensionMaster.objects.update_or_create(
                dimension_id=dimension_id,
                defaults=defaults,
            )
            if was_created:
                created += 1

        return {"created": created}

    def _seed_root_causes(self, workbook) -> dict[str, int]:
        worksheet = workbook[SHEET_ROOT_CAUSES]
        created = 0

        for row in worksheet.iter_rows(
            min_row=ROOT_CAUSE_DATA_START_ROW,
            values_only=True,
        ):
            code = self._cell_text(row, 0)
            if not code or not code.startswith("RC-"):
                continue

            name = self._cell_text(row, 1)
            description = self._cell_text(row, 2)
            standard_remediation_pattern = self._cell_text(row, 4)

            _, was_created = RootCauseMaster.objects.update_or_create(
                code=code,
                defaults={
                    "name": name,
                    "description": description,
                    "standard_remediation_pattern": standard_remediation_pattern,
                },
            )
            if was_created:
                created += 1

        return {"created": created}

    def _seed_checks(
        self,
        workbook,
        catalogue_by_check_id: dict[str, dict[str, str]],
    ) -> dict[str, int]:
        worksheet = workbook[SHEET_CHECKS]
        created = 0
        known_root_cause_codes = set(
            RootCauseMaster.objects.values_list("code", flat=True)
        )
        dimension_by_id = {
            item.dimension_id: item
            for item in DimensionMaster.objects.all()
        }
        optional_check_ids = OPTIONAL_CHECK_IDS

        for row in worksheet.iter_rows(
            min_row=CHECK_DATA_START_ROW,
            values_only=True,
        ):
            check_id = self._cell_text(row, 1)
            if not check_id:
                continue

            dimension_label = self._cell_text(row, 3)
            parsed = self._parse_dimension_label(dimension_label)
            if parsed is None:
                raise CommandError(
                    f"Check {check_id} has unrecognized dimension: {dimension_label!r}"
                )

            dimension_id, _name = parsed
            dimension = dimension_by_id.get(dimension_id)
            if dimension is None:
                raise CommandError(
                    f"Check {check_id} references unknown dimension {dimension_id}"
                )

            catalogue = catalogue_by_check_id.get(check_id)
            if catalogue is None:
                raise CommandError(
                    f"Check {check_id} not found in {SHEET_CATALOGUE}"
                )

            root_cause_codes = self._parse_root_cause_codes(catalogue["root_causes"])
            missing_codes = sorted(
                code for code in root_cause_codes if code not in known_root_cause_codes
            )
            if missing_codes:
                raise CommandError(
                    f"Check {check_id} references unknown root causes: "
                    f"{', '.join(missing_codes)}"
                )

            check, was_created = CheckMaster.objects.update_or_create(
                check_id=check_id,
                defaults={
                    "sequence": self._to_int(row[0]),
                    "check_name": self._cell_text(row, 4),
                    "dimension": dimension,
                    "check_class": self._cell_text(row, 2),
                    "check_type": self._cell_text(row, 5),
                    "role": self._cell_text(row, 7),
                    "cadence": self._cell_text(row, 8),
                    "phase": self._cell_text(row, 9),
                    "systems_compared": catalogue["systems_compared"],
                    "numeric_weight": self._to_int(row[6], default=0),
                    "severity": catalogue["severity"],
                    "root_cause_ids": root_cause_codes,
                    "suggested_fix": catalogue.get("suggested_fix") or "",
                    "fix_type": catalogue.get("fix_type") or "",
                    "fix_owner": catalogue.get("fix_owner") or "",
                    "is_active": True,
                    "is_optional": check_id in optional_check_ids,
                },
            )

            if was_created:
                created += 1

        return {"created": created}

    def _load_catalogue_rows(self, workbook) -> dict[str, dict[str, str]]:
        worksheet = workbook[SHEET_CATALOGUE]
        catalogue_by_check_id: dict[str, dict[str, str]] = {}

        for row in worksheet.iter_rows(
            min_row=CATALOGUE_DATA_START_ROW,
            values_only=True,
        ):
            check_id = self._cell_text(row, 0)
            if not check_id or "-" not in check_id:
                continue

            catalogue_by_check_id[check_id] = {
                "systems_compared": self._cell_text(row, 4),
                "root_causes": self._cell_text(row, 11),
                "severity": self._cell_text(row, 14),
                "suggested_fix": self._cell_text(row, 16),
                "fix_type": self._cell_text(row, 17),
                "fix_owner": self._cell_text(row, 18),
            }

        return catalogue_by_check_id

    def _collect_check_dimension_labels(self, workbook) -> set[str]:
        worksheet = workbook[SHEET_CHECKS]
        labels: set[str] = set()

        for row in worksheet.iter_rows(
            min_row=CHECK_DATA_START_ROW,
            values_only=True,
        ):
            label = self._cell_text(row, 3)
            if label:
                labels.add(label)

        return labels

    @staticmethod
    def _parse_dimension_label(label: str) -> tuple[str, str] | None:
        match = DIMENSION_LABEL_RE.match(label.strip())
        if not match:
            return None
        return match.group(1), match.group(2).strip()

    @staticmethod
    def _parse_root_cause_codes(value: str) -> list[str]:
        if not value:
            return []
        return ROOT_CAUSE_CODE_RE.findall(value)

    @staticmethod
    def _parse_percent_from_rule(rule: str) -> int | None:
        if not rule:
            return None
        match = PERCENT_IN_RULE_RE.search(rule)
        if match is None:
            return None
        return int(match.group(1))

    @staticmethod
    def _cell_text(row: tuple, index: int) -> str:
        if index >= len(row) or row[index] is None:
            return ""
        return str(row[index]).strip()

    @staticmethod
    def _to_int(value, *, default: int | None = None) -> int:
        if value is None or value == "":
            if default is None:
                raise ValueError("Expected numeric value")
            return default
        return int(float(value))

    @staticmethod
    def _to_optional_number(value):
        if value is None or value == "":
            return None
        if isinstance(value, bool):
            return value
        number = float(value)
        if number.is_integer():
            return int(number)
        return number
